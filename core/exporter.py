"""Exportacion basica a Excel (modo demo o busqueda real). Sin logica de
postulacion, sin historial -- solo vuelca lo que ya se muestra en
pantalla, con un formato prolijo (freeze, autofiltro, colores por
decision, links clickeables)."""
import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill

RESULTADOS_DIR = "resultados"

_COLORES_DECISION = {
    "POSTULAR": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "REVISAR": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "DESCARTAR": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}


def construir_ruta_excel(prefijo="demo"):
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    carpeta_mes = datetime.datetime.now().strftime("%Y-%m")
    base = Path(RESULTADOS_DIR) / carpeta_mes
    base.mkdir(parents=True, exist_ok=True)
    return base / f"{prefijo}_{timestamp}.xlsx"


def crear_hoja_resumen(writer, datos_resumen):
    wb = writer.book
    ws = wb.create_sheet("RESUMEN")

    es_real = datos_resumen.get("modo") == "real"
    ws["A1"] = f"Agente Laboral Gen — Resumen de la búsqueda {'real' if es_real else 'demo'}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")

    etiqueta_font = Font(bold=True)
    modo_texto = (
        "BÚSQUEDA REAL — ofertas obtenidas de Computrabajo en vivo"
        if es_real else
        "DEMO — datos simulados, sin conexión a portales reales"
    )
    campos = [
        ("Fecha/hora", datos_resumen["fecha_hora"]),
        ("Modo", modo_texto),
        ("Ofertas mostradas", datos_resumen["cant_resultados"]),
        ("POSTULAR", datos_resumen["postular"]),
        ("REVISAR", datos_resumen["revisar"]),
        ("DESCARTAR", datos_resumen["descartar"]),
    ]
    fila = 3
    for etiqueta, valor in campos:
        ws.cell(row=fila, column=1, value=etiqueta).font = etiqueta_font
        ws.cell(row=fila, column=2, value=valor)
        fila += 1

    fila += 1
    ws.cell(row=fila, column=1, value="Acciones sugeridas").font = etiqueta_font
    fila += 1
    conteo_acciones = datos_resumen.get("conteo_acciones") or {}
    for etiqueta in ("POSTULAR HOY", "REVISAR ANTES DE POSTULAR", "REVISAR MANUALMENTE", "NO ACCIONAR"):
        ws.cell(row=fila, column=1, value=etiqueta).font = etiqueta_font
        ws.cell(row=fila, column=2, value=conteo_acciones.get(etiqueta, 0))
        fila += 1

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 45


def _formatear_hojas(writer):
    wb = writer.book
    for nombre_hoja in wb.sheetnames:
        if nombre_hoja == "RESUMEN":
            continue
        ws = wb[nombre_hoja]
        if ws.max_row < 2:
            continue

        encabezados = [c.value for c in ws[1]]
        col_link_idx = encabezados.index("link") + 1 if "link" in encabezados else None
        col_decision_idx = (
            encabezados.index("decision_sugerida") + 1 if "decision_sugerida" in encabezados else None
        )

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for col_cells in ws.columns:
            largo = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
            letra = col_cells[0].column_letter
            ws.column_dimensions[letra].width = min(max(largo + 2, 10), 60)

        if col_link_idx:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_link_idx)
                if cell.value:
                    cell.hyperlink = cell.value
                    cell.value = "Abrir oferta"
                    cell.style = "Hyperlink"

        if col_decision_idx:
            for row in range(2, ws.max_row + 1):
                valor = ws.cell(row=row, column=col_decision_idx).value
                relleno = _COLORES_DECISION.get(str(valor))
                if relleno:
                    for c in ws[row]:
                        c.fill = relleno


def exportar_excel(nuevas, acciones_df, datos_resumen):
    """Escribe RESUMEN + RESULTADOS (+ ACCIONES si hay alguna) en
    resultados/YYYY-MM/<prefijo>_<timestamp>.xlsx -- prefijo "demo" o
    "busqueda_real" segun datos_resumen["modo"]. Devuelve la ruta."""
    prefijo = "busqueda_real" if datos_resumen.get("modo") == "real" else "demo"
    archivo = construir_ruta_excel(prefijo=prefijo)
    datos_resumen = dict(datos_resumen, archivo=str(archivo))

    with pd.ExcelWriter(archivo, engine="openpyxl") as writer:
        crear_hoja_resumen(writer, datos_resumen)
        nuevas.to_excel(writer, sheet_name="RESULTADOS", index=False)
        if not acciones_df.empty:
            acciones_df.to_excel(writer, sheet_name="ACCIONES", index=False)
        _formatear_hojas(writer)

        wb = writer.book
        if "RESUMEN" in wb.sheetnames:
            idx = wb.sheetnames.index("RESUMEN")
            if idx != 0:
                wb.move_sheet("RESUMEN", offset=-idx)
        if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1 and wb["Sheet"]["A1"].value is None:
            wb.remove(wb["Sheet"])

    return archivo
